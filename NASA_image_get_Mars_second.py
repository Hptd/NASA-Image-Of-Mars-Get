#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import openpyxl
import requests
import sys
from openpyxl import load_workbook
from bs4 import BeautifulSoup


class Nasa1(object):
    """
    下载nasa图片
    """
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        }
        # self.url = "https://mars.nasa.gov/mars2020/multimedia/images/"
        self.url = "https://www.jpl.nasa.gov/images"
        self.local_path = "I:/NASAImageGet/NASAimageGet/NASA_Image_Mars"
        self.excel_path = 'I:/NASAImageGet/NASAimageGet/NASA_Image_Mars/NASA_Image_Mars_Second.xlsx'
        # self.nasa_host = "https://mars.nasa.gov"
        self.nasa_host = "https://www.jpl.nasa.gov"

    def run(self):
        """
        获取图片列表
        :return:
        """
        jump_url_list = []
        page = 0    # 开始位置
        while True:
            page += 1
            if page <= 40:
                continue
            elif page == 61:
                break
            print("当前页数: %s" % (int(page)))
            params = {
                "page": page
            }
            res = ""
            try:
                res = requests.get(self.url, params, headers=self.headers)
            except Exception as e:
                print("get pic list Error: %s" % str(e))

            # 当获取列表异常时，结束
            if not res:
                return

            # 获取列表数据
            data = res.content
            soup = BeautifulSoup(data, "lxml")
            a = soup.select('div[class="relative mb-6"]')
            if not a:
                break

            # 遍历本页列表，获取跳转链接
            for b in a:
                c = BeautifulSoup((str(b)), "lxml").select("a")[0]
                jump_url_list.append(self.nasa_host + c["href"])

            # for list_test in jump_url_list:
            #     print(list_test)

        print("图片总数：%s" % len(jump_url_list))
        num = 0     # 开始下载的图片序号
        # 获取详情页信息
        for idx, info_url in enumerate(jump_url_list[num:]):
            file_name = num + idx + 1  # 本地图片名称
            print("获取第%s张图片" % file_name)

            title, intro, pic_down_url = self.get_info(info_url)

            # 判断存储文件路径是否存在，不存在则创建
            path = self.download_path(self.local_path)
            full_file_name = "%s/%s.jpg" % (path, file_name)
            if not self.download_pic(pic_down_url, full_file_name):
                continue

            # 写入excel
            self.write_excel([[file_name, title, intro]])

    def get_info(self, url):
        """
        获取详情页信息
        :param url:
        :return:
        """
        response = ""
        for i in range(5):
            try:
                response = requests.get(url, headers=self.headers, timeout=3)
            except:
                print('请求链接超时，第%s次重复请求' % i)
                if i == 10:
                    print('请求链接失败，请查看网络连接并重启程序进行重试。')
                    os.system('pause')
                    sys.exit()

        if not response:
            return

        data = response.content
        soup = BeautifulSoup(data, "lxml")
        # 获取标题
        title = soup.select('h1')[0].text.strip()
        # 获取简介
        a = soup.select('div[class="BlockText text-body-lg"]')
        b = BeautifulSoup(str(a[0]), "lxml")
        content_list = []
        for c in b.select('p'):
            content_list.append(c.text)

        # 获取下载地址
        # down_url = self.nasa_host + soup.select('a[class="BaseButton text-contrast-none w-full mb-5 -primary -compact inline-block"]')[0]["href"]
        down_url = soup.select('a[class="BaseButton text-contrast-none w-full mb-5 -primary -compact inline-block"]')[
                       0]["href"]

        return title, "\n".join(content_list), down_url

    def download_pic(self, url, file_name):
        """
        下载图片
        :param url:
        :param file_name:
        :return:
        """
        for i in range(10):
            try:
                req = requests.get(url, timeout=5, stream=True, verify=True, headers=self.headers)

                with open(file_name, "ab") as f:
                    for chunk in req.iter_content(chunk_size=1024):
                        if chunk:
                            f.write(chunk)
                return True

            except Exception as e:
                print('第%s次获取图片异常，err: %s' % (i, str(e)))
                if i == 10:
                    return False

    def write_excel(self, pic_info_list):
        """
        写入excel
        :param pic_info_list:
        :return:
        """
        max_row_num = 0
        # 判断文件是否存在
        if not os.path.exists(self.excel_path):
            pic_info_list.insert(0, ["图片ID", "图片标题", "图片说明"])
            workbook = openpyxl.Workbook()  # 新建excel
            workbook.create_sheet('Sheet1')
            sheet = workbook.active
        else:
            workbook = load_workbook(self.excel_path)   # 读取excel
            sheet = workbook.active
            max_row_num = sheet.max_row

        # 写入数据
        for idx, row in enumerate(pic_info_list):
            for idx1, cols in enumerate(row):
                sheet.cell(row=idx + max_row_num + 1, column=idx1 + 1, value=cols)

        workbook.save(self.excel_path)

    def download_path(self, path):
        """
        创建下载文件夹目录
        :param path:
        :return:
        """
        if not os.path.exists(path):
            os.makedirs(path)
        return path


if __name__ == '__main__':
    Nasa1().run()
