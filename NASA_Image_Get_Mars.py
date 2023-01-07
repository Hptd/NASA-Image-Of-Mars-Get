#!/usr/bin/env python
# -*- coding: utf-8 -*-
# import json
import os

import openpyxl
import requests
import sys


class Nasa(object):
    """
    下载nasa图片
    """
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.0.0 Safari/537.36',
        }
        self.url_main = 'https://www.nasa.gov/sites/default/files/'
        self.url = 'https://www.nasa.gov/api/2/ubernode/_search'
        # self.local_path = "/Users/renguangcheng/Desktop/nasa"
        # self.excel_path = '/Users/renguangcheng/Desktop/nasa.xlsx'
        self.local_path = "I:/NASAImageGet/NASAimageGet/NASA_Image_Mars"
        self.excel_path = "I:/NASAImageGet/NASAimageGet/NASA_Image_Mars/NASA_Image_Mars.xlsx"
    def run(self):
        """
        获取图片列表
        :return:
        """
        # 首次获取列表，获取总数
        page = 24
        params = {
            'size': str(page),
            'from': '0',
            'sort': 'promo-date-time:desc',
            'q': '((ubernode-type:image) AND (missions:3643))',
            '_source_include': 'promo-date-time,master-image,nid,title,topics,missions,collections,other-tags,ubernode-type,primary-tag,secondary-tag,cardfeed-title,type,collection-asset-link,link-or-attachment,pr-leader-sentence,image-feature-caption,attachments,uri',
        }
        res_nasa = self.get_index_html(self.url, params)
        if not res_nasa:
            print("未获取到数据")
            return
        json_nasa = res_nasa.json()

        # 获取图片总数
        total = json_nasa["hits"]["total"]
        # 获取全部图片
        params["size"] = total
        res_nasa = self.get_index_html(self.url, params)
        json_nasa = res_nasa.json()
        url_local_list = json_nasa['hits']['hits']

        # 遍历图片下载
        num = 0     # 开始下载的图片序号
        pic_info_list = []  # 图片详情列表
        for idx, pic in enumerate(url_local_list[num:]): # [num:9] 9就是需要爬取的数量，默认不填则认定爬取全部
            file_name = idx + 1  # 本地图片名称
            print("获取第%s张图片" % file_name)
            pic_local = pic['_source']['master-image']['uri']
            pic_url = self.url_main + pic_local[9:]
            print("full_url:----->", pic_url)

            # 判断存储文件路径是否存在，不存在则创建
            path = self.download_path(self.local_path)
            full_file_name = "%s/%s.jpg" % (path, file_name)
            if self.download_pic(pic_url, full_file_name):
                # 保存图片详情
                pic_info_list.append([file_name, pic['_source']["title"], pic['_source']["image-feature-caption"]])
                    # if pic_info_list:
                        # 写入excel
                self.write_excel(pic_info_list)

    def get_index_html(self, url, params):
        """
        获取主页面
        :param url:
        :param params:
        :return:
        """
        response = ""
        for i in range(5):
            try:
                response = requests.get(url, headers=self.headers, timeout=3, params=params)
            except:
                print('请求链接超时，第%s次重复请求' % i)
                if i == 4:
                    print('请求链接失败，请查看网络连接并重启程序进行重试。')
                    os.system('pause')
                    sys.exit()
        return response

    def download_pic(self, url, file_name):
        """
        下载图片
        :param url:
        :param file_name:
        :return:
        """
        for i in range(5):
            try:
                req = requests.get(url, timeout=5, stream=True, verify=True, headers=self.headers)

                with open(file_name, "ab") as f:
                    for chunk in req.iter_content(chunk_size=1024):
                        if chunk:
                            f.write(chunk)
                return True

            except Exception as e:
                print('第%s次获取图片异常，err: %s' % (i, str(e)))
                if i == 4:
                    return False

    def write_excel(self, pic_info_list):
        """
        写入excel
        :param pic_info_list:
        :return:
        """
        # pic_info_list.insert(0, ["图片ID", "图片标题", "图片说明"])
        workbook = openpyxl.Workbook()  # 新建excel
        workbook.create_sheet('Sheet1')

        sheet = workbook.active
        # 写入数据
        for idx, row in enumerate(pic_info_list):
            for idx1, cols in enumerate(row):
                sheet.cell(row=idx + 1, column=idx1 + 1, value=cols)

        workbook.save(self.excel_path)

    def download_path(self, path):
        """
        创建下载文件夹目录
        :param path:
        :return:
        """
        if not os.path.exists(path):
            os.makedirs(path)
        return path


if __name__ == '__main__':
    Nasa().run()
